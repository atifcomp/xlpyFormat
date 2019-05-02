# -*- coding: utf-8 -*-
"""
Created on Thu May  2 15:13:40 2019

@author: amomin
"""

from openpyxl.reader.excel import load_workbook
from openpyxl.styles.cell_style import CellStyle
from copy import copy, deepcopy

srcfile = 'C:/Users/amomin/Documents/GitHub/xlpyFormat/template.xlsx'
srcfile2 = 'C:/Users/amomin/Documents/GitHub/xlpyFormat/data.xlsx'
book = load_workbook(filename = srcfile)    
book2 = load_workbook(filename = srcfile2)

default_sheet = book['template_sheet']
default_sheet2 = book2['data_sheet']

#
#def copy_cell(cell, new_cell):
#    new_cell.value = cell.value
#    if cell.has_style:
#        new_cell._style = copy(cell._style)
##    return new_cell
#
#
#
#for row in default_sheet.rows:
#    for cell in row:
#        new_cell = default_sheet2.cell(row=cell.row,
#                   column=cell.column, value= cell.value)
#        if cell.has_style:
#            copy_cell(cell, new_cell)
#
#book2.save(filename=srcfile2)


def copy_style(src_cell, dest_cell):
    dest_cell.font = copy(src_cell.font)
    dest_cell.fill = copy(src_cell.fill)
    dest_cell.border = copy(src_cell.border)
    dest_cell.alignment = copy(src_cell.alignment)
    dest_cell.number_format = copy(src_cell.number_format)


for row in default_sheet.rows:
    for cell in row:
        new_cell = default_sheet2.cell(row=cell.row,
                   column=cell.column, value= cell.value)
        if cell.has_style:
            copy_style(cell, new_cell)
            

x = default_sheet.merged_cells
x        

book2.save(filename=srcfile2)
