import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.reader.excel import load_workbook
import re
from xlsxwriter.utility import xl_cell_to_rowcol



class FormatWriter():
    """
    it format excel report
    Parameters
    ----------
    srcfile : string, required
        the path of the excel file to format
        
    """
    
    def __init__(self, srcfile): 
        self.wb = load_workbook(filename = srcfile)
        self.srcfile = srcfile
    def save_workbook(self):
        self.wb.save(self.srcfile)
        
    def set_all_borders(self,sheetname,rng): 
        """
        this takes sheetname and range as input and apply all borders same as all borders of
        excel
        """        
        try:
            ws = self.wb[sheetname]
            startCell,endCell = re.split(':',rng.strip())
            minRow,minCol = xl_cell_to_rowcol(startCell)
            maxRow,maxCol = xl_cell_to_rowcol(endCell)
            
            border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

            rows = ws.iter_rows(min_row=minRow,max_row=maxRow,min_col=minCol,max_col=maxCol)
            for row in rows:
                for cell in row:                    
                    cell.border = border
        except:
            print("set_all_borders, sheetname or ranges not provided correctly")


    
################################################################################################
sourceFile = "C:/Users/amomin/Desktop/Projects/Charter ISP/Days Aging.xlsx"

#Calling part starts from here
formatter = FormatWriter(sourceFile)
formatter.set_all_borders("abc","A1:Z1")
formatter.save_workbook()