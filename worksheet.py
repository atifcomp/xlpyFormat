# openpyxl.__version__ = 2.6.2
# Created By: Atif Momin
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import get_column_letter
import re
from xlsxwriter.utility import xl_cell_to_rowcol
from copy import copy
#from openpyxl.styles.alignment import Alignment
import __formats



class Worksheet():
    """
    This is the class as worksheet level
    
    Parameters:    
    wb:     workbook object, required
            this is the object of work book class of xlpyformatter
    ws_name: string, required
        this is the worksheet name parameter                
    """
    
    def __init__(self,wb,ws_name):
        self.wb = wb
        self.ws = wb[ws_name]
        self.lastRow = self.ws.max_row
        self.lastCol = self.ws.max_column        


    def _sequence_check(self,firstCol,lastCol):
        """
        helper function which checks if the ranges provided are in correct order, return True or False
        ----------        
        """
        firstCol = int(firstCol.encode("utf-8").hex(),16)
        lastCol = int(lastCol.encode("utf-8").hex(),16)
        if firstCol>lastCol:
            return False
        else:
            return True
    
    def _letter_to_col_number(self,col_rng):
        """
        helper function converts the range of column to column number
        ----------        
        """
        firstCol,lastCol = re.split(':',col_rng.strip())
        _,num1 = xl_cell_to_rowcol(firstCol+'1')
        _,num2 = xl_cell_to_rowcol(lastCol+'1')
        return num1,num2
    
    def _alpha_check(self,col_rng):
        """
        helper function checks if the column range provided is alpha and valid column range
        ----------        
        """
        firstCol,lastCol = re.split(':',col_rng.strip())            
        
        if firstCol.isalpha() and lastCol.isalpha():
            return self._sequence_check(firstCol,lastCol)
        else:
            raise Exception('column provided not alpha')        
    
        
    
    def set_all_borders(self,rng=None,border_type=None): 
        """
        This takes sheetname and range as input and apply all borders same as all borders of
        excel
        ----------
        rng : string, required
             range of the excel where border has to be applied
        """        
        self._general_formating_range(rng,_set_all_borders=True,_border_type=border_type)
#        try:
#            startCell,endCell = re.split(':',rng.strip())
#            minRow,minCol = xl_cell_to_rowcol(startCell)
#            maxRow,maxCol = xl_cell_to_rowcol(endCell)
#            
#            border = Border(left=Side(border_style='thin', color='000000'),
#                right=Side(border_style='thin', color='000000'),
#                top=Side(border_style='thin', color='000000'),
#                bottom=Side(border_style='thin', color='000000'))
#
#            rows = self.ws.iter_rows(min_row=minRow+1,max_row=maxRow+1,min_col=minCol+1,max_col=maxCol+1)
#            for row in rows:
#                for cell in row:                    
#                    cell.border = border
#        except:
#            print("set_all_borders, sheetname or ranges not provided correctly")

    def _general_formating_range(self,rng,**kwargs):
        startCell,endCell = re.split(':',rng.strip())
        minRow,minCol = xl_cell_to_rowcol(startCell)
        maxRow,maxCol = xl_cell_to_rowcol(endCell)
        
        rows = self.ws.iter_rows(min_row=minRow+1,max_row=maxRow+1,min_col=minCol+1,max_col=maxCol+1)
        for row in rows:
            for cell in row:                    
                if '_set_all_borders' in kwargs:
                    border = Border(left=Side(border_style=kwargs['_border_type'], color='373636'),
                            right=Side(border_style=kwargs['_border_type'], color='373636'),
                            top=Side(border_style=kwargs['_border_type'], color='373636'),
                            bottom=Side(border_style=kwargs['_border_type'], color='373636'))                    
                    cell.border = border
                else:
                    pass


    def column_autofit(self):
        """
        This takes auto fits the columns of the sheet
        ----------        
        """
        print(self.ws.columns)
        for col in self.ws.columns:        
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                if cell.coordinate in self.ws.merged_cells: # not check merge_cells
                    continue
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass            
            adjusted_width = (max_length + 4)
            self.ws.column_dimensions[column].width = adjusted_width
    
    def column_width(self,col_name,col_width):
        self.ws.column_dimensions[col_name].width = col_width
                        
    def column_range_width(self,col_rng,col_width):        
        """
        This takes column range and column widht as input and set the width of columns
        ----------
        col_rng : string, required
             eg. 'A:Z'
        col_width : integer, required
             width to be set
        """        
        try:
            if self._alpha_check(col_rng):
                num1 , num2 = self._letter_to_col_number(col_rng)
                for _i in range(num1+1,num2+2):
                    self.ws.column_dimensions[get_column_letter(_i)].width = col_width            
            else:
                raise Exception('column range is not provided correctly')
        except:
            print("set_all_borders, sheetname or ranges not provided correctly")

    def column_set_format(self,col_rng,formatType):
        """
        This takes column range and format type as input and apply format to columns
        ----------
        col_rng : string, required
             eg. 'A:Z'
        formatType : string, required
             this is the format type as per openpyxl formats complete list is provided on below path
             https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html             
        """
        try:
            self._general_formating_col(col_rng,_formatType=formatType)
        except:
            print("set_format, colrg not provided correctly or failed some conversion")
    
    def column_center_align(self,col_rng):
        try:
            self._general_formating_col(col_rng,_column_center_align=True)
        except:
            print("column_center_align, colrg not provided correctly or failed some conversion")
    
    def column_apply_font(self,col_rng,font_format):
        _font_format = { 'name':None,
                'size':None,
                'bold':None,
                'italic':None,
                'vertAlign':None,
                'underline':None,
                'strike':None,
                'color':None}
        _font_format.update(font_format)
        print(_font_format)
        font = Font(name=_font_format['name'],
                size=_font_format['size'],
                bold=_font_format['bold'],
                italic=_font_format['italic'],
                vertAlign=_font_format['vertAlign'],
                underline=_font_format['underline'],
                strike=_font_format['strike'],
                color=_font_format['color'] )
        print(font)
        self._general_formating_col(col_rng,_font=font)
        
    
    def _general_formating_col(self,col_rng,**kwargs):         
        if self._alpha_check(col_rng):
            num1,num2 = self._letter_to_col_number(col_rng)            
            for _i in range(num1+1,num2+2):
                for _j in range(2,self.lastRow+1):
                    if '_formatType' in kwargs:
                        self.ws[str(get_column_letter(_i)+str(_j))].number_format = kwargs['_formatType']
                    elif '_column_center_align' in kwargs:
                        self.ws[str(get_column_letter(_i)+str(_j))].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif '_font' in kwargs:                        
                        self.ws[str(get_column_letter(_i)+str(_j))].font = kwargs['_font']                        
        else:
                raise Exception('column range is not provided correctly')
            
            
    
class Copyformat(Worksheet):
    """
    This class is for copying the formats of one worksheet to other, it inherites 
    the properties of worksheet class as we have to use few functionalities of worksheet 
    class such as column autofit.
    
    Parameters:    
    wb:     workbook object, required
            this is the object of work book class of xlpyformatter, this is initialize 
            from worksheet class
    ws_name: string, required
            this is the worksheet name parameter, this is initialize 
            from worksheet class
    wb_template: workbook object, required
            work book object of template excel
    ws_template_name: string, required
            worksheet nanme of the template 
            
    """
    def __init__(self,wb,ws_name,wb_template,ws_template_name):
        super().__init__(wb,ws_name)
        self.wb_template = wb_template
        self.ws_template = self.wb_template[ws_template_name]
    
    def copy_style(self,src_cell, dest_cell):
        dest_cell.font = copy(src_cell.font)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.border = copy(src_cell.border)
        dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.number_format = copy(src_cell.number_format)    
    
    
    def replicate_format(self):                
        # this code copies the formats of template sheet and apply it over to 
        # destinations sheet
        for row in self.ws_template.rows:
            for cell in row:
                new_cell = self.ws.cell(row=cell.row,
                           column=cell.column)
                if cell.has_style:
                    self.copy_style(cell, new_cell)
        
        #auto-fitting the column as it doesn't fit automatically by copying styles        
        #this function has been inherited from Worksheet (parent) class
        self.column_autofit()
        
        #if sheet has merged cells, it doesn't copy it by default
        #we need to merge it afterwards
        print(self.ws_template.merged_cells)
        if self.ws_template.merged_cells:
            for rng in self.ws_template.merged_cells:
                self.ws.merge_cells(str(rng))
    
            
        